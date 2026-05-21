"""
admin_utils.py — общие admin-команды для program_bot и lead_bot.

Доступные команды:
  /getxls  — выгружает все таблицы БД в Excel с реальными фото
  /getdb   — отправляет сырой .db файл
"""
import logging
import os
import sqlite3
import tempfile
from datetime import datetime as dt
from pathlib import Path

import database as db

logger = logging.getLogger(__name__)


# ── Excel-экспорт ─────────────────────────────────────────────

def export_db_to_excel(output_path: str):
    """Выгружает все основные таблицы БД в Excel (каждая — отдельный лист)."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    TABLES = [
        ("users",             "Пользователи"),
        ("task_completions",  "Задания"),
        ("user_achievements", "Ачивки"),
        ("leads",             "Лиды"),
    ]

    HEADER_FONT  = Font(bold=True, color="FFFFFF")
    HEADER_FILL  = PatternFill("solid", fgColor="2E4057")
    HEADER_ALIGN = Alignment(horizontal="center", vertical="center")

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    conn = sqlite3.connect(str(db.DB_PATH))
    conn.row_factory = sqlite3.Row
    try:
        # ── Обычные таблицы ───────────────────────────────────
        for table, sheet_title in TABLES:
            try:
                cursor = conn.execute(f"SELECT * FROM {table}")
                rows   = cursor.fetchall()
                cols   = [d[0] for d in cursor.description]
            except Exception as e:
                logger.warning(f"getxls: пропускаю таблицу {table}: {e}")
                continue

            ws = wb.create_sheet(title=sheet_title)
            ws.freeze_panes = "A2"
            ws.row_dimensions[1].height = 20

            for ci, col in enumerate(cols, 1):
                cell = ws.cell(row=1, column=ci, value=col)
                cell.font      = HEADER_FONT
                cell.fill      = HEADER_FILL
                cell.alignment = HEADER_ALIGN

            for ri, row in enumerate(rows, 2):
                for ci in range(1, len(cols) + 1):
                    ws.cell(row=ri, column=ci, value=row[ci - 1])

            for ci, col in enumerate(cols, 1):
                max_len = len(str(col))
                for row in rows:
                    v = row[ci - 1]
                    max_len = max(max_len, len(str(v)) if v is not None else 0)
                ws.column_dimensions[get_column_letter(ci)].width = min(max_len + 2, 40)

        # ── Лист с реальными фото ─────────────────────────────
        try:
            from openpyxl.drawing.image import Image as XlImage
            from PIL import Image as PilImage
            from io import BytesIO

            ws_ph = wb.create_sheet(title="Фото")
            ws_ph.freeze_panes = "A2"

            for ci, col in enumerate(["user_id", "photo_type", "created_at", "Фото"], 1):
                cell = ws_ph.cell(row=1, column=ci, value=col)
                cell.font      = HEADER_FONT
                cell.fill      = HEADER_FILL
                cell.alignment = HEADER_ALIGN

            ws_ph.column_dimensions["A"].width = 14
            ws_ph.column_dimensions["B"].width = 12
            ws_ph.column_dimensions["C"].width = 20
            ws_ph.column_dimensions["D"].width = 22

            photo_rows = conn.execute(
                "SELECT user_id, photo_type, created_at, photo_path, photo_data "
                "FROM user_photos ORDER BY user_id, id"
            ).fetchall()

            for ri, row in enumerate(photo_rows, 2):
                ws_ph.cell(row=ri, column=1, value=row[0])
                ws_ph.cell(row=ri, column=2, value=row[1])
                ws_ph.cell(row=ri, column=3, value=row[2])

                photo_path = row[3]
                photo_data = row[4]

                # Читаем из файла на диске (приоритет), затем из BLOB
                try:
                    if photo_path:
                        from pathlib import Path as _Path
                        _p = _Path(photo_path)
                        img = PilImage.open(_p) if _p.exists() else (
                            PilImage.open(BytesIO(photo_data)) if photo_data else None
                        )
                    elif photo_data:
                        img = PilImage.open(BytesIO(photo_data))
                    else:
                        img = None

                    if img is None:
                        ws_ph.cell(row=ri, column=4, value="нет данных")
                        continue

                    img.thumbnail((160, 160))
                    buf = BytesIO()
                    img.save(buf, format="JPEG")
                    buf.seek(0)
                    xl_img = XlImage(buf)
                    xl_img.anchor = f"D{ri}"
                    ws_ph.add_image(xl_img)
                    ws_ph.row_dimensions[ri].height = 122
                except Exception as e:
                    ws_ph.cell(row=ri, column=4, value=f"ошибка: {e}")

        except Exception as e:
            logger.warning(f"getxls: лист Фото не создан: {e}")

    finally:
        conn.close()

    wb.save(output_path)


# ── Telegram-хендлеры ─────────────────────────────────────────

def make_admin_commands(admin_id: int):
    """
    Возвращает готовые async-хендлеры cmd_getxls и cmd_getdb,
    привязанные к конкретному admin_id.
    """
    from telegram.ext import ContextTypes
    from telegram import Update

    async def cmd_getxls(update: Update, context: ContextTypes.DEFAULT_TYPE):
        """Выгружает БД в Excel и отправляет файл (только для admin)."""
        if update.effective_user.id != admin_id:
            return
        msg = await update.message.reply_text("⏳ Формирую Excel...")
        tmp_path = None
        try:
            tmp_fd, tmp_path = tempfile.mkstemp(suffix=".xlsx")
            os.close(tmp_fd)
            export_db_to_excel(tmp_path)
            size_kb = os.path.getsize(tmp_path) // 1024
            with open(tmp_path, "rb") as f:
                await update.message.reply_document(
                    document=f,
                    filename=f"zarik_{dt.now().strftime('%Y%m%d_%H%M')}.xlsx",
                    caption=f"📊 База данных · {size_kb} КБ · 5 листов",
                )
            await msg.delete()
        except Exception as e:
            logger.exception("cmd_getxls error")
            await msg.edit_text(f"❌ Ошибка: {e}")
        finally:
            if tmp_path and os.path.exists(tmp_path):
                os.unlink(tmp_path)

    async def cmd_getdb(update: Update, context: ContextTypes.DEFAULT_TYPE):
        """Отправляет сырой .db файл (только для admin)."""
        if update.effective_user.id != admin_id:
            return
        db_path = db.DB_PATH
        if not Path(db_path).exists():
            await update.message.reply_text("❌ Файл базы не найден.")
            return
        size_kb = Path(db_path).stat().st_size // 1024
        await update.message.reply_document(
            document=open(db_path, "rb"),
            filename=f"zarik_{dt.now().strftime('%Y%m%d_%H%M')}.db",
            caption=f"🗄 База данных · {size_kb} КБ",
        )

    return cmd_getxls, cmd_getdb
